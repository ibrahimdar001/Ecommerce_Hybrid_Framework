import openpyxl

def getRowCount(file, sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return (sheet.max_row)

def getColumnCount(file, sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return (sheet.max_column)

def ReadData(file, sheetname, rowNo, colNo):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return (sheet.cell(row=rowNo, column=colNo).value)

def WriteData(file, sheetname, rowNo, colNo, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    sheet.cell(row=rowNo, column=colNo).value = data
    workbook.save(file)

